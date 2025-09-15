from openpyxl import Workbook,load_workbook

#from new_page.a import excel_dosya

wb = load_workbook("veriler.xlsx")
ws = wb.active
for satir in range(1,ws.max_row+1):
    for sutun in range(1,ws.max_column+1):
        print(" | " + str(ws.cell(satir,sutun).value) + " | ",end="")
    print()
#print(ws["A7"].value)